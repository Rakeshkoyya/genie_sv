"""File parsing service for extracting text from various file formats."""

import io
from typing import Any
from pypdf import PdfReader
from openpyxl import load_workbook


async def parse_pdf(content: bytes) -> tuple[str, dict[str, Any]]:
    """Extract text from a PDF file.
    
    Args:
        content: PDF file bytes
        
    Returns:
        Tuple of (extracted_text, metadata)
    """
    reader = PdfReader(io.BytesIO(content))
    text_parts = []
    
    for page in reader.pages:
        page_text = page.extract_text()
        if page_text:
            text_parts.append(page_text)
    
    extracted_text = "\n\n".join(text_parts)
    metadata = {
        "page_count": len(reader.pages),
    }
    
    return extracted_text, metadata


async def parse_excel(content: bytes, filename: str = "") -> tuple[str, dict[str, Any]]:
    """Extract text from an Excel file.
    
    Args:
        content: Excel file bytes
        filename: Original filename (for extension detection)
        
    Returns:
        Tuple of (extracted_text, metadata)
    """
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    text_parts = []
    sheet_names = []
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_names.append(sheet_name)
        text_parts.append(f"## Sheet: {sheet_name}")
        
        for row in sheet.iter_rows(values_only=True):
            row_values = [str(cell) if cell is not None else "" for cell in row]
            if any(row_values):  # Skip entirely empty rows
                text_parts.append("\t".join(row_values))
        
        text_parts.append("")  # Empty line between sheets
    
    workbook.close()
    
    extracted_text = "\n".join(text_parts)
    metadata = {
        "sheet_count": len(sheet_names),
        "sheet_names": sheet_names,
    }
    
    return extracted_text, metadata


async def parse_csv(content: bytes) -> tuple[str, dict[str, Any]]:
    """Extract text from a CSV file.
    
    Args:
        content: CSV file bytes
        
    Returns:
        Tuple of (extracted_text, metadata)
    """
    import csv
    
    text = content.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    
    # Convert to tab-separated for consistency with Excel output
    extracted_text = "\n".join("\t".join(row) for row in rows)
    metadata = {
        "row_count": len(rows),
        "column_count": len(rows[0]) if rows else 0,
    }
    
    return extracted_text, metadata


async def parse_text(content: bytes) -> tuple[str, dict[str, Any]]:
    """Extract text from a plain text file.
    
    Args:
        content: Text file bytes
        
    Returns:
        Tuple of (extracted_text, metadata)
    """
    # Try UTF-8 first, fall back to latin-1
    try:
        text = content.decode("utf-8")
    except UnicodeDecodeError:
        text = content.decode("latin-1", errors="replace")
    
    lines = text.split("\n")
    metadata = {
        "line_count": len(lines),
        "char_count": len(text),
    }
    
    return text, metadata


def get_file_type(filename: str, mime_type: str = "") -> str:
    """Determine the file type from filename and MIME type.
    
    Args:
        filename: Original filename
        mime_type: MIME type if available
        
    Returns:
        File type string: 'pdf', 'image', 'excel', 'csv', 'text', 'document', 'other'
    """
    ext = filename.split(".")[-1].lower() if "." in filename else ""
    
    if ext == "pdf" or mime_type == "application/pdf":
        return "pdf"
    
    if ext in ("png", "jpg", "jpeg", "gif", "webp") or mime_type.startswith("image/"):
        return "image"
    
    if ext in ("xlsx", "xls") or mime_type in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel"
    ):
        return "excel"
    
    if ext == "csv" or mime_type == "text/csv":
        return "csv"
    
    if ext in ("txt", "md", "json", "xml", "html", "htm", "rtf", "log") or mime_type.startswith("text/"):
        return "text"
    
    if ext in ("doc", "docx", "ppt", "pptx", "odt", "ods", "odp"):
        return "document"
    
    return "other"


class FileParserService:
    """Service class for file parsing operations."""
    
    async def parse(
        self,
        content: bytes,
        filename: str,
        mime_type: str = ""
    ) -> tuple[str | None, dict[str, Any]]:
        """Parse a file and extract text content.
        
        Args:
            content: File bytes
            filename: Original filename
            mime_type: MIME type if available
            
        Returns:
            Tuple of (extracted_text or None, metadata)
        """
        file_type = get_file_type(filename, mime_type)
        metadata = {"mimeType": mime_type, "fileType": file_type}
        
        if file_type == "pdf":
            text, pdf_meta = await parse_pdf(content)
            metadata.update(pdf_meta)
            return text, metadata
        
        elif file_type == "excel":
            text, excel_meta = await parse_excel(content, filename)
            metadata.update(excel_meta)
            return text, metadata
        
        elif file_type == "csv":
            text, csv_meta = await parse_csv(content)
            metadata.update(csv_meta)
            return text, metadata
        
        elif file_type == "text":
            text, text_meta = await parse_text(content)
            metadata.update(text_meta)
            return text, metadata
        
        # Images and documents: no text extraction, store as-is
        elif file_type in ("image", "document", "other"):
            return None, metadata
        
        return None, metadata
